"""Genera catalogo.json — el catálogo maestro de formularios PLOG.

Fuente de verdad: Excel `Definicion_Politicas_Formularios_PLOG_v3 Zenput.xlsx`
(políticas por zona) + mapeo de form_template_ids verificado contra el API en vivo
(discovery 2026-08-11) + correcciones de Roberto 2026-08-11:

  1. PRO-SUC-1: solo las versiones EN USO por PLOG (954599 "2025V1" + 954592 "1.1");
     901109 excluido (es la variante que usa CAS, 2/40 submissions PLOG).
  2. Conteo Diario de Pollo (1108481): FUERA del sistema.
  3. PRO-CSC-5 Revisión a Vehículos (954594): DENTRO.
  4. CHECKLIST AUDITORÍA PROVEEDORES (1695114): DENTRO (calificación, 8 áreas;
     política propuesta mensual — validar con directores).

Correr:  python3 etl_plog/config/build_catalogo.py [ruta_excel]
"""
import json
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
EXCEL_DEFAULT = "/Volumes/aduanadata/Definicion_Politicas_Formularios_PLOG_v3 Zenput.xlsx"
ZONAS_HOJAS = {
    "PLOG Nuevo León": "nuevo_leon",
    "PLOG Laguna": "laguna",
    "PLOG Querétaro": "queretaro",
}

# familia -> (clave de fila en el Excel, form_template_ids verificados en API)
# Las series traen dict {ft: sufijo} para cumplimiento por día/semana.
FAMILIAS = {
    "alistamiento_a": {
        "excel_key": "954598 y serie",
        "fts": {954598: "lun", 1040520: "mar", 1040521: "mie", 1040522: "jue",
                1040523: "vie", 1040524: "sab", 1010538: "dom"},
        "serie": "dia_semana",
    },
    "alistamiento_l": {
        "excel_key": "1040507 y serie",
        "fts": {1040507: "lun", 1040506: "mar", 1040504: "mie", 1040505: "jue",
                1038660: "vie", 1038659: "sab", 1038658: "dom"},
        "serie": "dia_semana",
    },
    "rl1_entrega_t1": {"excel_key": "954595", "fts": [954595]},
    "rl2_entrega_t2": {"excel_key": "954602", "fts": [954602]},
    "checklist_cierre": {"excel_key": "1040519", "fts": [1040519]},
    "alistamiento_servir": {"excel_key": "1040698", "fts": [1040698]},
    "alistamiento_hornos": {"excel_key": "997146", "fts": [997146]},
    "deposito_valores": {
        "excel_key": "997145/1043394/1043393",
        # ft distinto por zona (regional)
        "fts_por_zona": {"nuevo_leon": [997145], "laguna": [1043394], "queretaro": [1043393]},
    },
    "autogestion_calidad": {
        "excel_key": "1560060 y serie",
        "fts": {1560060: "s1", 1568530: "s2", 1568532: "s3", 1568536: "s4"},
        "serie": "semana_del_mes",
    },
    "pro_suc_4_autoeval": {"excel_key": "954596", "fts": [954596]},
    "mtto_mensual": {"excel_key": "1572636", "fts": [1572636]},
    "mtto_trimestral": {"excel_key": "1572200", "fts": [1572200]},
    "mtto_semestral": {"excel_key": "1572620", "fts": [1572620]},
    "visita_negocio": {"excel_key": "1059179", "fts": [1059179]},
    "visita_seguimiento": {"excel_key": "954591", "fts": [954591]},
    "rh1_visita": {"excel_key": "954604", "fts": [954604]},
    # Corrección Roberto: versiones EN USO (901109 = variante CAS, excluida)
    "pro_suc_1": {"excel_key": "901109/954592", "fts": [954599, 954592]},
    "pro_suc_3": {"excel_key": "1038657/1059183", "fts": [1038657, 1059183]},
    "do_supervision_operativa": {"excel_key": "1161748", "fts": [1161748]},
    "do_control_seguridad": {"excel_key": "1161749", "fts": [1161749]},
    "vcal_calidad_integral": {"excel_key": "954593/1059169", "fts": [954593, 1059169]},
    "comisariato_evaluacion": {"excel_key": "954601 y rel.", "fts": [954601, 954600, 1059170, 1059182]},
    "recorrido_comisariato": {"excel_key": "1161752", "fts": [1161752]},
    "pro_sc_6_seguridad": {"excel_key": "954603", "fts": [954603]},
    "pro_csc_5_vehiculos": {"excel_key": "954594", "fts": [954594]},
    "revision_operativa_plog": {"excel_key": "1638930", "fts": [1638930]},
    "gestion_finanzas_plog": {"excel_key": "1638942", "fts": [1638942]},
    "matriz_imagen_plog": {"excel_key": "1657858", "fts": [1657858]},
    # No estaba en el Excel — alta directa (corrección Roberto)
    "auditoria_proveedores": {"excel_key": None, "fts": [1695114]},
}

# Patrón de extracción de score por familia (discovery 2026-08-11):
#   p1 = formula "PORCENTAJE %" / "CALIFICACION PORCENTAJE %" (estilo CAS)
#   p2 = formula "... (100%)" / "... Puntuación (100%)"
#   p3 = formula "Calificación General ..." + áreas "AREA <X>" (mantenimientos)
#   p4 = "SUBTOTAL - <SECCIÓN>" + "RESULTADO TOTAL ..." (secciones ponderadas)
SCORE_PATRON = {
    "alistamiento_servir": "p2", "alistamiento_hornos": "p2",
    "autogestion_calidad": "p1", "pro_suc_4_autoeval": "p2",
    "mtto_mensual": "p3", "mtto_trimestral": "p3", "mtto_semestral": "p3",
    "pro_suc_1": "p2", "pro_suc_3": "p2",
    "do_supervision_operativa": "p1", "do_control_seguridad": "p1",
    "vcal_calidad_integral": "p2", "comisariato_evaluacion": "p2",
    "pro_sc_6_seguridad": "p2",
    "revision_operativa_plog": "p4", "gestion_finanzas_plog": "p4",
    "auditoria_proveedores": "p3",  # 9 formulas + sección "Calificacion General"; confirmar con 1a submission
}

# Overrides puntuales (decisiones Roberto / hallazgos discovery).
OVERRIDES = {
    "1108481": {"excluir": True, "motivo": "Roberto 2026-08-11: Conteo de Pollo no cuenta"},
    "954594": {"forzar_usar": True, "motivo": "Roberto 2026-08-11: Vehículos sí se tiene en cuenta"},
}

ALTAS_DIRECTAS = {
    "auditoria_proveedores": {
        "nombre": "CHECKLIST AUDITORÍA PROVEEDORES PLOG",
        "medicion": "Calificación",
        "frecuencia_propuesta": "Mensual",
        "nota": "Alta directa Roberto 2026-08-11 (no está en el Excel v3). 0 submissions aún; "
                "8 áreas: Instalaciones, Higiene Personal, Manejo de Alimentos, Temperatura, "
                "Limpieza, Plagas, Almacenamiento, Maquila. Política por validar con directores.",
        "zonas": {z: {"usar": "Sí", "medicion": "Calificación", "frecuencia": "Mensual",
                      "hora_limite": "", "dias_gracia": "0", "sucursales_aplica": "Por definir"}
                  for z in ("nuevo_leon", "laguna", "queretaro")},
    },
}


def _norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return s.strip().lower()


def leer_excel(path):
    wb = load_workbook(path, data_only=True)
    politicas = {}
    for hoja, zkey in ZONAS_HOJAS.items():
        for r in list(wb[hoja].iter_rows(values_only=True))[2:]:
            if not (r[0] and r[1]) or str(r[0]).strip() == "Form ID":
                continue
            key = str(r[0]).strip()
            e = politicas.setdefault(key, {"nombre": str(r[1]).strip(), "zonas": {}})
            usar = _norm(r[6])
            e["zonas"][zkey] = {
                "usar": "Sí" if usar.startswith("si") else ("No" if usar.startswith("no") else str(r[6] or "?")),
                "medicion": str(r[7] or "?").strip(),
                "frecuencia": str(r[8] or "?").strip(),
                "horario_esperado": str(r[9] or "").strip(),
                "hora_limite": str(r[10] or "").strip(),
                "dias_gracia": str(r[11] or "0").strip(),
                "agenda": str(r[12] or "").strip(),
                "calificacion_vista": str(r[15] or "").strip(),
                "sucursales_aplica": (str(r[17]).strip() if len(r) > 17 and r[17] else ""),
            }
    return politicas


def construir(excel_path):
    politicas = leer_excel(excel_path)
    catalogo = {"_generado": "build_catalogo.py — NO editar a mano; corre el script",
                "_fuente_excel": str(excel_path), "familias": {}}
    for fam, spec in FAMILIAS.items():
        ek = spec.get("excel_key")
        if ek and ek in OVERRIDES and OVERRIDES[ek].get("excluir"):
            continue
        if ek is None:
            alta = ALTAS_DIRECTAS[fam]
            entrada = {"nombre": alta["nombre"], "zonas": alta["zonas"], "nota": alta["nota"]}
        else:
            pol = politicas.get(ek)
            if pol is None:
                raise SystemExit(f"familia {fam}: clave Excel {ek!r} no encontrada")
            entrada = {"nombre": pol["nombre"], "zonas": pol["zonas"]}
            if ek in OVERRIDES and OVERRIDES[ek].get("forzar_usar"):
                for z in entrada["zonas"].values():
                    if not _norm(z["usar"]).startswith("si"):
                        z["usar"] = "Sí"
                        z["nota_override"] = OVERRIDES[ek]["motivo"]
        if "fts_por_zona" in spec:
            entrada["form_templates_por_zona"] = spec["fts_por_zona"]
        elif isinstance(spec["fts"], dict):
            entrada["form_templates"] = spec["fts"]
            entrada["serie"] = spec["serie"]
        else:
            entrada["form_templates"] = spec["fts"]
        entrada["score_patron"] = SCORE_PATRON.get(fam)
        # tipo consolidado: calificación si alguna zona lo pide con score
        medic = {_norm(z.get("medicion")) for z in entrada["zonas"].values()}
        if any(m.startswith("calific") for m in medic) and any(m.startswith(("cumplim", "ambos")) for m in medic):
            entrada["tipo"] = "ambos"
        elif any(m.startswith("ambos") for m in medic):
            entrada["tipo"] = "ambos"
        elif any(m.startswith("calific") for m in medic):
            entrada["tipo"] = "calificacion"
        else:
            entrada["tipo"] = "cumplimiento"
        # el Excel dice "Por definir" en los forms nuevos, pero tienen score verificado
        if entrada["score_patron"] and entrada["tipo"] == "cumplimiento":
            entrada["tipo"] = "ambos"
        catalogo["familias"][fam] = entrada
    return catalogo


if __name__ == "__main__":
    excel = sys.argv[1] if len(sys.argv) > 1 else EXCEL_DEFAULT
    cat = construir(excel)
    out = HERE / "catalogo.json"
    out.write_text(json.dumps(cat, ensure_ascii=False, indent=1))
    fams = cat["familias"]
    print(f"catalogo.json: {len(fams)} familias")
    for tipo in ("cumplimiento", "calificacion", "ambos"):
        n = [f for f, e in fams.items() if e["tipo"] == tipo]
        print(f"  {tipo}: {len(n)} → {', '.join(n)}")
